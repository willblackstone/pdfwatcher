#!/usr/bin/env python3

import os
import time
import re
from datetime import datetime
import fitz  # PyMuPDF
import PySimpleGUI as sg
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side, Font

# Styles and defaults
DEFAULT_PAGE_PREFIX = "CHR"
DEFAULT_ROW_HEADING = "Daily Notes"
COLUMN_WIDTH_INFO = 40  # ≈11.7 cm
FILL_COLOR = "CADBEA"
BORDER_STYLE = Side(style='thin')

# Patterns for parsing
date_pattern = re.compile(r'^\d{1,2} [A-Za-z]{3} \d{4}$')
time_pattern = re.compile(r'^\d{2}:\d{2}$')

class PDFChangeHandler(FileSystemEventHandler):
    def __init__(self, pdf_path, xlsx_path, prefix, row_heading, mode, join_order, window):
        self.pdf_path = pdf_path
        self.xlsx_path = xlsx_path
        self.prefix = prefix
        self.row_heading = row_heading
        self.mode = mode  # 'Comment → Time' or 'Time → Comment'
        self.join_order = join_order  # 'Top to Bottom' or 'Bottom to Top'
        self.window = window
        self.seen = set()

    def on_modified(self, event):
        if event.src_path == self.pdf_path:
            raw = extract_highlighted_text(self.pdf_path)
            records = parse_records(raw, self.prefix, self.row_heading, self.mode, self.join_order)
            new = [r for r in records if r not in self.seen]
            if new:
                append_to_xlsx(self.xlsx_path, new)
                self.seen.update(new)
                for rec in new:
                    timestamp = datetime.now().strftime('%H:%M:%S')
                    msg = f"{timestamp} Appended {rec[1]}: '{rec[2]}'"
                    self.window.write_event_value('-LOG-', msg)


def extract_highlighted_text(pdf_path):
    highlights = []
    doc = fitz.open(pdf_path)
    for page_index in range(len(doc)):
        page = doc[page_index]
        for annot in page.annots() or []:
            if annot.type[0] == 8:
                rects = []
                for i in range(0, len(annot.vertices), 4):
                    quad = annot.vertices[i:i+4]
                    xs = [pt[0] for pt in quad]
                    ys = [pt[1] for pt in quad]
                    rects.append(fitz.Rect(min(xs), min(ys), max(xs), max(ys)))
                words = page.get_text('words')
                picked = [(w[3], w[0], w[4]) for w in words if any(r.intersects(fitz.Rect(w[:4])) for r in rects)]
                if picked:
                    picked.sort()
                    text = ' '.join(w[2] for w in picked).strip()
                    highlights.append((page_index + 1, text))
    return highlights


def parse_records(highlights, prefix, row_heading, mode, join_order):
    records = []
    current_date = None
    comment_buffer = []
    pending_time = None

    def flush_group(comments, time_str, time_page):
        # if user chooses Bottom to Top, reverse list
        if join_order == 'Bottom to Top':
            comments = list(reversed(comments))
        joined = ' [...] '.join(comments)
        page_str = f"{prefix} {time_page}"
        info = f"{row_heading}\n{time_str}: {joined}"
        date_str = current_date or datetime.now().strftime('%d/%m/%Y')
        records.append((date_str, page_str, info))

    for page, text in highlights:
        text = text.strip()
        if date_pattern.match(text):
            dt = datetime.strptime(text, '%d %b %Y')
            current_date = dt.strftime('%d/%m/%Y')
            comment_buffer.clear()
            pending_time = None
        elif time_pattern.match(text):
            time_str = text
            if mode == 'Comment → Time':
                if comment_buffer:
                    comments = [c for (_, c) in comment_buffer]
                    flush_group(comments, time_str, page)
                    comment_buffer.clear()
            else:  # Time → Comment
                if pending_time and comment_buffer:
                    flush_group(comment_buffer, pending_time, page)
                    comment_buffer.clear()
                pending_time = time_str
        else:
            c = re.sub(r',\s*$', ' [...]', text.replace('\n', ' '))
            if mode == 'Comment → Time':
                comment_buffer.append((page, c))
            else:
                comment_buffer.append(c)
    # flush tail for Time → Comment
    if mode == 'Time → Comment' and pending_time and comment_buffer:
        flush_group(comment_buffer, pending_time, page)
    return records


def append_to_xlsx(xlsx_path, rows):
    if not os.path.exists(xlsx_path):
        wb = Workbook()
        ws = wb.active
        ws.append(['Date', 'Page Number', 'Info'])
        for cell in ws[1]:
            cell.border = Border(left=BORDER_STYLE, right=BORDER_STYLE,
                                 top=BORDER_STYLE, bottom=BORDER_STYLE)
    else:
        wb = load_workbook(xlsx_path)
        ws = wb.active

    ws.column_dimensions['C'].width = COLUMN_WIDTH_INFO
    fill = PatternFill(start_color=FILL_COLOR, end_color=FILL_COLOR, fill_type='solid')
    thin = Border(left=BORDER_STYLE, right=BORDER_STYLE,
                  top=BORDER_STYLE, bottom=BORDER_STYLE)

    for date_str, page_str, info in rows:
        ws.append([date_str, page_str, info])

    prev_date = None
    band = False
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3):
        date_cell, _, info_cell = row
        if date_cell.value != prev_date:
            prev_date = date_cell.value
            band = not band
        if band:
            for cell in row:
                cell.fill = fill
        for cell in row:
            cell.border = thin
        info_cell.alignment = Alignment(wrap_text=True)
        info_cell.font = Font(bold=False)

    wb.save(xlsx_path)


def start_watcher(values, window):
    handler = PDFChangeHandler(
        values['-PDF-'], values['-XLSX-'],
        values['-PREFIX-'] or DEFAULT_PAGE_PREFIX,
        values['-HEADING-'] or DEFAULT_ROW_HEADING,
        values['-ORDER-'], values['-JOIN-'], window)
    obs = Observer()
    obs.schedule(handler, path=os.path.dirname(values['-PDF-']), recursive=False)
    obs.start()
    return obs


def main():
    sg.theme('SystemDefault')
    layout = [
        [sg.Text('Page Prefix:'), sg.Input(DEFAULT_PAGE_PREFIX, key='-PREFIX-', size=(10,1)),
         sg.Text('Row Heading:'), sg.Input(DEFAULT_ROW_HEADING, key='-HEADING-', size=(15,1))],
        [sg.Text('Highlight Order:'), sg.Combo(['Comment → Time', 'Time → Comment'], default_value='Comment → Time', key='-ORDER-'),
         sg.Text('Join Order:'), sg.Combo(['Top to Bottom', 'Bottom to Top'], default_value='Top to Bottom', key='-JOIN-')],
        [sg.FileBrowse('Select PDF…', file_types=(('PDF Files','*.pdf'),), key='-PDFB-'), sg.Input(key='-PDF-', enable_events=True, size=(40,1))],
        [sg.FileBrowse('Select Excel…', file_types=(('Excel Files','*.xlsx'),), key='-XLSXB-'), sg.Input(key='-XLSX-', enable_events=True, size=(40,1))],
        [sg.Button('Start Watching', key='-START-'), sg.Button('Stop Watching', key='-STOP-', disabled=True)],
        [sg.Text('Status:'), sg.Text('Stopped', key='-STATUS-')],
        [sg.Text('Log:')], [sg.Multiline(key='-LOG-', size=(80,10), autoscroll=True, disabled=True)]
    ]
    window = sg.Window('PDF Watcher', layout, finalize=True)
    observer = None
    while True:
        event, values = window.read(timeout=200)
        if event in (sg.WIN_CLOSED, 'Exit'):
            break
        if event == '-START-':
            if not values['-PDF-'] or not values['-XLSX-']:
                sg.popup_error('Select PDF and Excel')
                continue
            observer = start_watcher(values, window)
            window['-STATUS-'].update('Watching')
            window['-START-'].update(disabled=True)
            window['-STOP-'].update(disabled=False)
        if event == '-STOP-':
            if observer:
                observer.stop()
                observer.join()
                observer = None
            window['-STATUS-'].update('Stopped')
            window['-START-'].update(disabled=False)
            window['-STOP-'].update(disabled=True)
        if event == '-LOG-':
            window['-LOG-'].print(values['-LOG-'])
    if observer:
        observer.stop()
        observer.join()
    window.close()

if __name__ == '__main__':
    main()
